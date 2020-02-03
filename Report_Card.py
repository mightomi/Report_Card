from datetime import datetime, date, timedelta
import inline as inline
import numpy as np
import pandas as pd
import psutil
import openpyxl
import requests
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import time
import os
import imgkit
import smtplib
from fpdf import FPDF
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


class Main:
    def __init__(self):
        self.cwd = os.getcwd()
        self.directory(self.cwd + "\\" + "Data")
        self.cloning_dataset("data\\" + 'Student Gradebook.xlsx',
                             "https://docs.google.com/spreadsheets/d/1HYjfEe3aCbufbqIXKs0Xz-gfoQNztGhCN1ivx0gZXnc/export?format = xlsx")
        self.cloning_dataset("data\\" + 'users.csv',
                             "https://docs.google.com/spreadsheets/d/1hTKuo8BCw3wuIt7ITiFjxEE9Xs-Gb9HIPo9GUSis-Vc/export?format=csv")
        data_full = self.extract_data("Data\\" + "Student Gradebook.xlsx", "overall")
        data_month = self.extract_data("Data\\" + "Student Gradebook.xlsx", "month")
        self.user_data = self.get_user_data()
        data_month = self.data_cleaning(data_month)
        data_full = self.data_cleaning(data_full)
        self.directory(self.cwd + "\\" + "Report_card")

        month_now = pd.Timestamp.now().month
        if month_now == 1:
            month_now = 13
        months = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October",
                  "November", "December"]
        self.given_month = months[month_now - 1]
        self.directory(self.cwd + "\\" + "Report_card\\" + self.given_month)
        df = data_month
        new_df = pd.DataFrame()
        for j in range(len(df['id'].drop_duplicates().values)):
            count = 0
            for i in range(len(df)):
                if df['id'].drop_duplicates().values[j] == df['id'].values[i]:
                    count += 1
            if count > 2:
                new_df = new_df.append(df[df['id'] == df['id'].drop_duplicates().values[j]])
        df = new_df

        data_topper = new_df.groupby('Student')['Points'].sum().sort_values(ascending=False).reset_index()

        self.user_data['lname'].fillna("", inplace=True)

        for i in range(len(self.user_data["fname"].values)):
            given_name = self.user_data["fname"][i]
            last_name = self.user_data["lname"][i]
            self.directory(
                self.cwd + "\\" + "Report_card\\" + self.given_month + "\\" + given_name + "_" + last_name)
            self.file_loc = self.cwd + "\\" + "Report_card\\" + self.given_month + "\\" + given_name + "_" + last_name + "\\"
            self.radar_plot(data_month, given_name)

            data_topper = new_df.groupby('Student')['Points'].sum().sort_values(ascending=False).reset_index()
            task_topper = new_df.groupby("Task")["Points"].max().reset_index()
            new_df = self.comparison_plot(df, new_df, given_name)
            data = pd.merge(new_df, task_topper, on='Task')
            data.rename(columns={'Points_x': 'Marks obtained', 'Points_y': 'Highest marks', 'Total': 'Total marks'},
                        inplace=True)
            df1 = self.report(given_name, data)
            self.table(df1, self.given_month)
            percent = self.percentage_cal(df1)
            rank = self.rank_cal(data_topper, given_name)
            student_data = data_full[data_full['Student'] == given_name]
            overall_percent = self.overall_percentage(student_data)
            overall_rank = self.overall_ranking(data_full, given_name)
            month_now_graph = pd.Timestamp.now().month
            main_data = pd.DataFrame()
            cats = []
            for j in range(12):
                data_month_graph = self.extract_data_for_monthly_graph(self.cwd + "\\Data\\" + "Student Gradebook.xlsx",
                                                                       j)
                if len(data_month_graph) == 0:
                    break
                data_month_graph = self.data_cleaning(data_month_graph)
                month = data_month_graph["Date_month"][0]
                if month == month_now_graph:
                    break
                months = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September",
                          "October", "November", "December"]
                month_check = months[month]
                year_check = str(data_month_graph["Date_year"][0])
                cats.append(month_check + ", " + year_check)
                data = self.calculate_performance_for_monthlygraph(data_month_graph, month_check, year_check,
                                                                   given_name)
                main_data = main_data.append(data, ignore_index=True)
            main_data['Month'] = pd.Categorical(main_data['Month'], ordered=True, categories=cats)
            self.monthly_graph(main_data)
            self.PDF_maker(given_name, last_name, self.given_month, percent, overall_percent, rank, overall_rank,
                           self.file_loc)

            email_send = self.user_data["email"][i]
            self.Mail_sender(given_name + "_" + last_name, email_send, overall_rank)
            print("Email successfully sent to " + email_send)

    def directory(self, path):
        try:
            os.mkdir(path)
        except:
            print("Folder already exists")

    def get_user_data(self):
        return pd.read_csv(self.cwd + "\\Data\\" + "users.csv")

    def cloning_dataset(self, path, url):
        resp = requests.get(url)
        output = open(path, 'wb')
        output.write(resp.content)
        output.close()

    def extract_data(self, file, condition):
        wb = openpyxl.load_workbook(file)
        data = pd.DataFrame()
        for i in range(len(wb.sheetnames)):
            try:
                if condition == 'overall':
                    data = pd.concat([data, pd.read_excel(file, sheet_name=i)])
                elif condition == 'month':
                    data = pd.read_excel(file, sheet_name=i)
            except:
                data = pd.read_excel(file, sheet_name=0)
        return data

    def data_cleaning(self, data):
        mask = data["Student"] != "Nitish"  # removing sir
        data = data[mask]
        data["id"] = pd.to_numeric(data["id"])
        data["Student"].replace("Swaastik", "Swaastick", inplace=True)
        data.reset_index(inplace=True, drop=True)  # giving index to data set
        data.dropna(inplace=True)
        data = data.iloc[:, :10]
        data["Student"].replace("Shakib", "Md Shakib", inplace=True)
        data["Student"].replace("Kunal", "Kunal N.", inplace=True)
        data["Student"].replace("Siddhishikha", "Sushree", inplace=True)
        data["Date"] = pd.to_datetime(data["Date"])
        data["Date_day"] = data["Date"].dt.day
        data["Date_month"] = data["Date"].dt.month
        data["Date_year"] = data["Date"].dt.year
        data.drop(columns=["Date"], inplace=True)
        return data

    def radar_plot(self, data, student_name):
        student_names = data["Student"].unique()
        arr_type_name = ["Consistency", "Curiosity", "Diligence", "Creativity", "Intent", "Sincerity", "Knowledge",
                         "Patience", "Analytical Ability", "Communication", "Confidence", "Hard-work"]
        type_cons_max = type_cur_max = type_dil_max = type_cre_max = type_int_max = type_sin_max = type_kno_max = type_pat_max = type_ana_max = type_comm_max = type_conf_max = type_hard_max = 0
        ctype_cons = ctype_cur = ctype_dil = ctype_cre = ctype_int = ctype_sin = ctype_kno = ctype_pat = ctype_ana = ctype_comm = ctype_conf = ctype_hard = 0

        for j in range(student_names.size):
            mask = data["Student"] == student_names[j]
            df = data[mask]
            df.reset_index(drop=True, inplace=True)
            type_cons = type_cur = type_dil = type_cre = type_int = type_sin = type_kno = type_pat = type_ana = type_comm = type_conf = type_hard = 0
            for i in range(df.shape[0]):
                type_cons += df["Type"][i].count("Consistency")
                type_cur += df["Type"][i].count("Curiosity")
                type_dil += df["Type"][i].count("Diligence")
                type_cre += df["Type"][i].count("Creativity")
                type_int += df["Type"][i].count("Intent")
                type_sin += df["Type"][i].count("sincerity")
                type_sin += df["Type"][i].count("Sincerity")
                type_kno += df["Type"][i].count("Knowledge")
                type_pat += df["Type"][i].count("Patience")
                type_ana += df["Type"][i].count("Analytical Ability")
                type_ana += df["Type"][i].count("Analysis")
                type_comm += df["Type"][i].count("Communication")
                type_comm += df["Type"][i].count("communication")
                type_conf += df["Type"][i].count("Confidence")
                type_hard += df["Type"][i].count("Hard-work")

                ctype_cons += df["Type"][i].count("Consistency")
                ctype_cur += df["Type"][i].count("Curiosity")
                ctype_dil += df["Type"][i].count("Diligence")
                ctype_cre += df["Type"][i].count("Creativity")
                ctype_int += df["Type"][i].count("Intent")
                ctype_sin += df["Type"][i].count("sincerity")
                ctype_sin += df["Type"][i].count("Sincerity")
                ctype_kno += df["Type"][i].count("Knowledge")
                ctype_pat += df["Type"][i].count("Patience")
                ctype_ana += df["Type"][i].count("Analytical Ability")
                ctype_ana += df["Type"][i].count("Analysis")
                ctype_comm += df["Type"][i].count("Communication")
                ctype_comm += df["Type"][i].count("communication")
                ctype_conf += df["Type"][i].count("Confidence")
                ctype_hard += df["Type"][i].count("Hard-work")

                type_cons_max = max(type_cons_max, type_cons)
                type_cur_max = max(type_cur_max, type_cur)
                type_dil_max = max(type_dil_max, type_dil)
                type_cre_max = max(type_cre_max, type_cre)
                type_int_max = max(type_int_max, type_int)
                type_sin_max = max(type_sin_max, type_sin)
                type_kno_max = max(type_kno_max, type_kno)
                type_pat_max = max(type_pat_max, type_pat)
                type_ana_max = max(type_ana_max, type_ana)
                type_comm_max = max(type_comm_max, type_comm)
                type_conf_max = max(type_conf_max, type_conf)
                type_hard_max = max(type_hard_max, type_hard)

        carr_type_max = [ctype_cons, ctype_cur, ctype_dil, ctype_cre, ctype_int, ctype_sin, ctype_kno, ctype_pat,
                         ctype_ana, ctype_comm, ctype_conf, ctype_hard]
        arr_type_max = [type_cons_max, type_cur_max, type_dil_max, type_cre_max, type_int_max, type_sin_max,
                        type_kno_max, type_pat_max, type_ana_max, type_comm_max, type_conf_max, type_hard_max]

        for i in range(len(arr_type_name)):
            if carr_type_max[i] < len(student_names):
                arr_type_max[i] = 0

        mask = data["Student"] == student_name
        df = data[mask]
        df.reset_index(drop=True, inplace=True)
        new_arr_type = []
        new_arr_type_name = []
        new_arr_type_max = []
        type_cons = type_cur = type_dil = type_cre = type_int = type_sin = type_kno = type_pat = type_ana = type_comm = type_conf = type_hard = 0
        for i in range(df.shape[0]):
            type_cons += df["Type"][i].count("Consistency") * df["Points"][i] / df["Total"][i]
            type_cur += df["Type"][i].count("Curiosity") * df["Points"][i] / df["Total"][i]
            type_dil += df["Type"][i].count("Diligence") * df["Points"][i] / df["Total"][i]
            type_cre += df["Type"][i].count("Creativity") * df["Points"][i] / df["Total"][i]
            type_int += df["Type"][i].count("Intent") * df["Points"][i] / df["Total"][i]
            type_sin += df["Type"][i].count("sincerity") * df["Points"][i] / df["Total"][i]
            type_sin += df["Type"][i].count("Sincerity") * df["Points"][i] / df["Total"][i]
            type_kno += df["Type"][i].count("Knowledge") * df["Points"][i] / df["Total"][i]
            type_pat += df["Type"][i].count("Patience") * df["Points"][i] / df["Total"][i]
            type_ana += df["Type"][i].count("Analytical Ability") * df["Points"][i] / df["Total"][i]
            type_ana += df["Type"][i].count("Analysis") * df["Points"][i] / df["Total"][i]
            type_comm += df["Type"][i].count("Communication") * df["Points"][i] / df["Total"][i]
            type_comm += df["Type"][i].count("communication") * df["Points"][i] / df["Total"][i]
            type_conf += df["Type"][i].count("Confidence") * df["Points"][i] / df["Total"][i]
            type_hard += df["Type"][i].count("Hard-work") * df["Points"][i] / df["Total"][i]

        arr_type = [type_cons, type_cur, type_dil, type_cre, type_int, type_sin, type_kno, type_pat, type_ana,
                    type_comm, type_conf, type_hard]

        for i in range(len(arr_type)):
            if (arr_type[i] != 0) & (arr_type_max[i] != 0):
                new_arr_type.append(arr_type[i])
                new_arr_type_name.append(arr_type_name[i])
                new_arr_type_max.append(arr_type_max[i])

        arr_type_percentage = []
        for i in range(len(new_arr_type_max)):
            arr_type_percentage.append(new_arr_type[i] / new_arr_type_max[i] * 10)

        df = pd.DataFrame(dict(r=arr_type_percentage, theta=new_arr_type_name))
        fig = px.line_polar(df, r='r', theta='theta', line_close=True)
        fig.update_traces(fill='toself')
        fig.write_image(self.file_loc + "radar_plot.jpg")

    def comparison_plot(self, df, new_df, student_name):
        mask = df["id"] != 0
        df = df[mask]
        mask = df["Student"] == student_name
        df_student = df[mask]
        topper_points = df.groupby("id")["Points"].max()
        total_points = df.groupby("id")["Total"].max()
        avg_points = df.groupby("id")["Points"].sum()
        total_avg_points = df.groupby("id")["Total"].sum()
        student_points = df_student.groupby("id")["Points"].max()

        student_percentage = student_points / total_points * 100
        topper_percentage = topper_points / total_points * 100
        avg_percentage = avg_points / total_avg_points * 100
        topper_percentage = topper_percentage.reset_index()
        topper_percentage.rename(columns={"id": "id", 0: "Topper's Percentage"}, inplace=True)
        student_percentage = student_percentage.reset_index()
        student_percentage.rename(columns={"id": "id", 0: "Student's Percentage"}, inplace=True)
        avg_percentage = avg_percentage.reset_index()
        avg_percentage.rename(columns={"id": "id", 0: "Average Percentage"}, inplace=True)

        percentage = student_percentage.merge(topper_percentage, on='id')
        percentage = percentage.merge(avg_percentage, on='id')
        percentage.fillna(0, inplace=True)

        for i in range(len(percentage)):
            elements = []
            if percentage["Student's Percentage"][i] == 0:
                for j in range(len(new_df)):
                    if new_df['id'][j] == percentage['id'][i]:
                        elements = [new_df['Task'][j], new_df['Module'][j], new_df['Type'][j], new_df['Total'][j],
                                    new_df['Date_day'][j], new_df['Date_month'][j], new_df['Date_year'][j]]
                        break
                new_df1 = {'id': percentage['id'][i], 'Task': elements[0], 'Module': elements[1], 'Type': elements[2],
                           'Student': student_name, 'Late Submission': 'NaN', 'Points': 0, 'Total': elements[3],
                           'Task Winner': 'NaN', 'Date_day': elements[4], 'Date_month': elements[5],
                           'Date_year': elements[6]}
                new_df = new_df.append(new_df1, ignore_index=True)

        task = df.groupby('id')['Task'].max().reset_index()
        percentage = percentage.merge(task, on='id')
        percentage.rename(columns={'Task_x': 'Task'}, inplace=True)
        percentage.drop(columns=['id'], inplace=True)
        percentage.set_index('Task').plot(kind='barh', grid=True)

        plt.savefig(self.file_loc + "comparison.jpg", dpi=300, bbox_inches='tight')
        plt.clf()

        return new_df

    def report(self, student_name, data):
        daf = data.loc[data['Student'] == student_name]
        df1 = daf.groupby('Task')['Marks obtained'].sum()
        df2 = daf.groupby('Task')['Highest marks'].sum()
        df3 = daf.groupby('Task')['Total marks'].sum()
        df4 = df1 / df3 * 100
        df1 = df1.reset_index()
        df2 = df2.reset_index()
        df3 = df3.reset_index()
        df4 = df4.reset_index()
        df1.rename(columns={0: "Marks obtained"}, inplace=True)
        df2.rename(columns={0: "Highest marks"}, inplace=True)
        df3.rename(columns={0: "Total marks"}, inplace=True)
        df4.rename(columns={0: "Percentage"}, inplace=True)
        df1 = pd.merge(df1, df2, on='Task')
        df1 = pd.merge(df1, df3, on='Task')
        df1 = pd.merge(df1, df4, on='Task')
        myday_student = daf[daf['Task'] == 'AjKyaUkhada'].drop_duplicates(subset=['Date_day'], keep='first')
        myday_highest = data[data['Task'] == 'AjKyaUkhada']
        myday_highest.drop_duplicates(subset=['Date_day'], keep='first', inplace=True)
        a = daf.groupby('Task')['Late Submission'].sum().reset_index()
        b = daf.groupby('Task')['Task Winner'].sum().reset_index()
        df1 = pd.merge(df1, a, on='Task')
        df1 = pd.merge(df1, b, on='Task')
        df1['Late Submission'].replace([0,1,2],['No','Yes','Yes'],inplace=True)
        df1['Task Winner'].replace([0,1,2],['No','Yes','Yes'],inplace=True)
        index = 0
        for i in range(len(data['Task'].unique())):
            if data['Task'].unique()[i] == 'AjKyaUkhada':
                index = i
                break
        df1["Marks obtained"][index] = len(myday_student)
        df1["Highest marks"][index] = len(myday_highest)
        df1["Total marks"][index] = len(myday_highest)
        df1["Percentage"][index] = len(myday_student) / len(myday_highest) * 100

        for i in range(len(df1)):
            if (df1['Task'][i] == 'AjKyaUkhada') | (df1['Task'][i] == 'Knowledge Sharing'):
                df1['Late Submission'][i] = 'Not Assigned'
                df1['Task Winner'][i] = 'Not Assigned'
            if df1["Marks obtained"][i] == 0:
                df1['Late Submission'][i] = 'Not Assigned'
                df1['Task Winner'][i] = 'Not Assigned'

        return df1

    def table(self, data, month):
        # Set CSS properties for th elements in dataframe
        th_props = [
            ('font-size', '25px'),
            ('text-align', 'center'),
            ('font-weight', 'bold'),
            ('color', '#0A0202'),
            ('background-color', '87ceeb')
        ]

        # Set CSS properties for td elements in dataframe
        td_props = [
            ('font-size', '15px'),
            ('text-align', 'left'),
            ('color', '#000000')

        ]

        # Set table styles
        styles = [
            dict(selector="th", props=th_props),
            dict(selector="td", props=td_props)
        ]

        cm = sns.light_palette("orange", as_cmap=True)
        cap = 'This is your performance for the month of ' + month + '.'

        data = (data.style
                .set_properties(**{'background-color': 'orange', 'border-color': 'white'})
                .background_gradient(cmap=cm)
                .highlight_max(subset=["Percentage"], color='yellow')
                .set_caption(cap)
                .set_table_styles(styles))

        path_wkthmltoimage = r'C:\Program Files\wkhtmltox\bin\wkhtmltoimage.exe'
        config = imgkit.config(wkhtmltoimage=path_wkthmltoimage)
        imgkit.from_string(data.render(), self.file_loc + "styled_table.jpg", config=config)

    def percentage_cal(self, data):
        grand_total = data["Marks obtained"].sum()
        total = data["Total marks"].sum()
        percent = (grand_total / total) * 100
        return percent

    def rank_cal(self, data, student_name):
        rank_month = data.index[data['Student'] == student_name][0] + 1
        return rank_month

    def overall_percentage(self, data):
        overall_marks = data.groupby('Student')['Points'].sum()
        total_overall = data.groupby('Student')['Total'].sum()
        overall_percent = overall_marks / total_overall * 100
        return overall_percent.values[0]

    def overall_ranking(self, data, student_name):
        overall = data.groupby('Student')['Points'].sum().sort_values(ascending=False)
        overall_rank_names = overall.reset_index()
        overall_rank = overall_rank_names.index[overall_rank_names['Student'] == student_name][0] + 1
        return overall_rank

    def extract_data_for_monthly_graph(self, file, i):
        wb = openpyxl.load_workbook(file)
        data = pd.DataFrame()
        if i < len(wb.sheetnames):
            try:
                data = pd.read_excel(file, sheet_name=i)
            except:
                data = pd.read_excel(file, sheet_name=0)
        return data

    def calculate_performance_for_monthlygraph(self, df, month, year, student_name):
        new_df = pd.DataFrame()
        for j in range(len(df['id'].drop_duplicates().values)):
            count = 0
            for i in range(len(df)):
                if (df['id'].drop_duplicates().values[j] == df['id'].values[i]):
                    count += 1
            if (count > 2):
                new_df = new_df.append(df[df['id'] == df['id'].drop_duplicates().values[j]])
        df = new_df
        mask = df["Student"] == student_name
        df_student = df[mask]
        topper_points = df.groupby("id")["Points"].max()
        total_points = df.groupby("id")["Total"].max()
        avg_points = df.groupby("id")["Points"].sum()
        total_avg_points = df.groupby("id")["Total"].sum()
        student_points = df_student.groupby("id")["Points"].max()

        student_percentage = student_points / total_points * 100
        topper_percentage = topper_points / total_points * 100
        avg_percentage = avg_points / total_avg_points * 100
        topper_percentage = topper_percentage.reset_index()
        topper_percentage.rename(columns={"id": "id", 0: "Topper's Percentage"}, inplace=True)
        student_percentage = student_percentage.reset_index()
        student_percentage.rename(columns={"id": "id", 0: "Student's Percentage"}, inplace=True)
        avg_percentage = avg_percentage.reset_index()
        avg_percentage.rename(columns={"id": "id", 0: "Average Percentage"}, inplace=True)

        percentage = student_percentage.merge(topper_percentage, on='id')
        percentage = percentage.merge(avg_percentage, on='id')
        percentage.fillna(0, inplace=True)
        percentage.drop(columns=["id"], inplace=True)

        student_percentage_month = percentage["Student's Percentage"].values.mean()
        topper_percentage_month = percentage["Topper's Percentage"].values.mean()
        avg_percentage_month = percentage["Average Percentage"].values.mean()

        data = {"Month": month + ", " + year, "Student's Performance": student_percentage_month,
                "Topper's Performance": topper_percentage_month, "Average Performance": avg_percentage_month}
        return data

    def monthly_graph(self, data):
        data = data.set_index('Month')
        sns.set()
        fig = sns.lineplot(data=data, sort=False, markers={"Student's Performance": 'o', "Topper's Performance": 'X',
                                                           "Average Performance": 'D'}, palette='prism')
        fig = fig.get_figure()
        fig.savefig(self.file_loc + "monthlygraph.jpg", dpi=300, bbox_inches="tight")
        plt.clf()

    def PDF_maker(self, fname, lname, month, percent_month, percent, rank_month, overall_rank, file_location):
        percent = round(percent)
        percent_month = round(percent_month)

        pdf = FPDF('P', 'mm', 'A4')
        pdf.add_page()

        pdf.set_fill_color(200, 220, 255)

        pdf.image(self.cwd + "\\Other_necessary_items\\" + "campusX_logo.png", x=10, y=5, w=20)
        if percent >= 92:
            pdf.image(self.cwd + "\\Other_necessary_items\\" + "grade-aa.jpg", x=168, y=2, w=27)
        elif (percent >= 84) & (percent < 92):
            pdf.image(self.cwd + "\\Other_necessary_items\\" + "grade-a.jpg", x=168, y=2, w=27)
        elif (percent >= 75) & (percent < 84):
            pdf.image(self.cwd + "\\Other_necessary_items\\" + "grade-b.jpg", x=168, y=2, w=27)
        elif (percent >= 60) & (percent < 75):
            pdf.image(self.cwd + "\\Other_necessary_items\\" + "grade-c.jpg", x=168, y=2, w=27)
        else:
            pdf.image(self.cwd + "\\Other_necessary_items\\" + "grade-d.jpg", x=168, y=2, w=27)

        pdf.set_font("times", 'B', size=50)
        pdf.cell(0, 10, txt="Report Card", ln=1, align="C")
        pdf.cell(0, 20, ln=2)

        pdf.set_font("times", 'B', size=18)
        pdf.cell(0, 7, txt="Name :- " + fname + " " + lname, ln=2, align="L")

        if overall_rank == 1:
            pdf.cell(0, 7, txt="Overall Rank :- " + str(overall_rank) + "st", ln=2, align="L")
        elif overall_rank == 3:
            pdf.cell(0, 7, txt="Overall Rank :- " + str(overall_rank) + "nd", ln=2, align="L")
        elif overall_rank == 4:
            pdf.cell(0, 7, txt="Overall Rank :- " + str(overall_rank) + "rd", ln=2, align="L")
        else:
            pdf.cell(0, 7, txt="Overall Rank :- " + str(overall_rank) + "th", ln=2, align="L")

        pdf.cell(0, 7, txt="Overall Percentage :- " + str(percent) + "%", ln=2, align="L")
        pdf.cell(0, 7, ln=2)

        pdf.set_font("times", 'B', size=18)
        pdf.cell(0, 7, txt="Your Stats for " + month + " is as follows :- ", ln=2, align="L")

        pdf.image(self.cwd + "\\Student_images\\" + fname + "_" + lname + ".jpg", x=158, y=35, w=40)

        pdf.image(self.cwd + "\\Report_card\\" + month + "\\" + fname + "_" + lname + "\\styled_table.jpg", x=5,
                  y=77, w=200)

        pdf.set_font("times", 'I', size=11)
        pdf.cell(0, 73, ln=2)

        pdf.set_font("times", 'B', size=16)
        pdf.cell(0, 7, ln=2)
        pdf.cell(0, 7, txt="Graphs showing your performance throughout this month are given below :-", border=0, ln=1,
                 align="L")

        pdf.set_font("times", 'BU', size=14)
        pdf.image(self.cwd + "\\Report_card\\" + month + "\\" + fname + "_" + lname + "\\radar_plot.jpg", x=36,
                  y=164, w=135)

        pdf.cell(0, 107, ln=2)
        pdf.cell(0, 5, txt="These skills were tested this month and here is your performance.", align='C', ln=1)
        pdf.cell(0, 5, txt="Your vs the topper vs average performance of each task is given below.", align='C', ln=1)
        pdf.image(self.cwd + "\\Report_card\\" + month + "\\" + fname + "_" + lname + "\\comparison.jpg", x=38,
                  y=18, w=135)

        pdf.cell(0, 108, ln=2)
        pdf.cell(0, 5, txt="Now, where do you stand compared to previous months.", align='C', ln=1)
        pdf.image(self.cwd + "\\Report_card\\" + month + "\\" + fname + "_" + lname + "\\monthlygraph.jpg", x=40,
                  y=130, w=130)

        pdf.cell(0, 93, ln=2)
        pdf.cell(0, 6, ln=2)
        pdf.set_font("times", 'B', size=17)
        if rank_month == 1:
            pdf.set_font("times", 'B', size=16)
            pdf.cell(0, 5, txt="Congo! You have achieved 1st position this month with a percentage of " + str(
                percent_month) + "%.", border=0, ln=2, align="U")
        elif rank_month == 2:
            pdf.set_font("times", 'B', size=16)
            pdf.cell(0, 5, txt="Congo! You have achieved 2nd position this month with a percentage of " + str(
                percent_month) + "%.", border=0, ln=2, align="U")
        elif rank_month == 3:
            pdf.set_font("times", 'B', size=16)
            pdf.cell(0, 5, txt="Congo! You have achieved 3rd position this month with a percentage of " + str(
                percent_month) + "%.", border=0, ln=2, align="U")
        elif (rank_month == 4) | (rank_month == 5):
            pdf.cell(0, 5,
                     txt="Well Done! You are in top 5 and have ranked " + str(rank_month) + "th this month with a",
                     border=0, ln=2, align="U")
            pdf.cell(0, 5, txt="percentage of " + str(percent_month) + "%.", border=0, ln=2, align="U")
        elif (rank_month >= 6) & (rank_month <= 10):
            pdf.cell(0, 5,
                     txt="Well Done! You are in top 10 and have ranked " + str(rank_month) + "th this month with a",
                     border=0, ln=2, align="U")
            pdf.cell(0, 5, txt="percentage of " + str(percent_month) + "%.", border=0, ln=2, align="U")
        else:
            pdf.cell(0, 5, txt="Your have ranked " + str(rank_month) + "th this month with a percentage of " + str(
                percent_month) + "% and your", border=0, ln=2, align="U")
            pdf.cell(0, 5, txt="performance needs to improve.", border=0, ln=2, align="U")

        pdf.cell(0, 5, ln=2)
        pdf.set_font("times", 'I', size=14)
        pdf.cell(0, 5, txt="N.B.:- The tasks which have been performed by other students and evaluated this month are",
                 ln=2, align='U')
        pdf.cell(0, 5, txt="           considered above. Previous tasks are not taken into consideration.", ln=2,
                 align='U')

        pdf.cell(0, 8, ln=1)
        pdf.set_font("times", 'B', size=18)
        pdf.cell(0, 8, ln=2, txt="Date :- " + str(
            date.today()) + "                                                             Place :- Kolkata", align='L')
        pdf.cell(0, 5, ln=1)
        pdf.set_font("times", 'I', size=12)
        pdf.cell(0, 2, ln=2,
                 txt="**This is a computer generated document. If you have any queries, please contact +91 8420166148.")

        pdf.output(file_location + "" + fname + "_" + lname + ".pdf")

    def Mail_sender(self, name, email_send, rank):
        email_user = "testreport2019@gmail.com"
        email_password = "abcd@1234"

        subject = "Monthly Report"
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['Subject'] = subject
        msg['To'] = email_send

        body1 = """

        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <meta http-equiv="X-UA-Compatible" content="ie=edge">
            <title>Document</title>
            <style>
            .div1{
                border: 2px solid blue;

                margin: 5px;
            }
            .div2{
                margin: 10px;
                border: 2px solid black;
                border-radius: 20px;
            }
            .div3{
                margin: 10px;
                background-color: rgba(240,230,240,.9);
                border-radius: 20px;
                box-shadow: 0 2px 2px 0 rgb(0 ,0,0, .6 ) , 0 0 0 1px rgba(0,0,0,.3);
            }
            .imgdiv{
                padding-top: 2px;;
                /* float: left; */
                text-align: center;
            }
            .logo{
                height: 100%;
                width: 80%;
                box-shadow:0 2px 2px 0 rgb(0 ,0,0, .4 ) , 0 0 0 1px rgba(0,0,0,.2);
            }

            .dot{
                height: 10px;
                width: 10px;
                background-color: black;
                border-radius: 50%;
                display: inline-block;
                margin-right: 4px;
            }
            p{
                /* margin-left:20px; */
                text-align: center;
                font-size: 18px;
                font-family: 'Lucida Sans', 'Lucida Sans Regular', 'Lucida Grande', 'Lucida Sans Unicode', Geneva, Verdana, sans-serif;
            }
            .arrow{
                height: 3%;
                width: 3%;
                margin-left: 48.5%;
                margin-top: 2px;
                margin-bottom: 2px;
            }
            </style>
        </head>
        <body>
            <div class="div1">


                <div class="div2">
                    <div class="imgdiv">
                        <img  class="logo" src="https://drive.google.com/uc?authuser=0&id=1P0JiUm01B1itolWCpu06xmHilQcVtTFb&export=download" alt="logo">
                    </div>
                        <div class="div3">
                            <br>
                            <p><span class="dot"></span>Things work out best for those who utilise each opportunity.</p>
                            <p><span class="dot"></span>You did it. And we are proud of you.</p>
                            <p><span class="dot"></span>Keep moving forward. Don't look back.</p>
                            <p><span class="dot"></span>Wish you all the best for your upcoming tasks. <b>ROCK IT!!</b></p>
                            <p><span class="dot"></span>Your result for the month is as follows:</p> 
                            <img class="arrow" src="https://drive.google.com/uc?authuser=0&id=1V0u4yQmIJ9f1HTPMHH6rLGx04cycefw-&export=download" alt="down">
                            <br>

                        </div>
                </div>
            </div>
        </body>
        </html>


        """

        body2 = """
        
        <html lang="en">
        <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="ie=edge">
        <title>Document</title>
        <style>
                .div1{
                    border: 2px solid blue;
                    
                    margin: 5px;
                }
                .div2{
                    margin: 10px;
                    border: 2px solid black;
                    border-radius: 20px;
                }
                .div3{
                    margin: 10px;
                    background-color: rgba(240,230,240,.9);
                    border-radius: 20px;
                    box-shadow: 0 2px 2px 0 rgb(0 ,0,0, .6 ) , 0 0 0 1px rgba(0,0,0,.3);
                }
                .imgdiv{
                    padding-top: 2px;;
                    /* float: left; */
                    text-align: center;
                }
                .logo{
                    height: 100%;
                    width: 80%;
                    box-shadow:0 2px 2px 0 rgb(0 ,0,0, .4 ) , 0 0 0 1px rgba(0,0,0,.2);
                }
            
                .dot{
                    height: 10px;
                    width: 10px;
                    background-color: black;
                    border-radius: 50%;
                    display: inline-block;
                    margin-right: 4px;
                }
                p{
                    /* margin-left:20px; */
                    text-align: center;
                    font-size: 18px;
                    font-family: 'Lucida Sans', 'Lucida Sans Regular', 'Lucida Grande', 'Lucida Sans Unicode', Geneva, Verdana, sans-serif;
                }
                .arrow{
                    height: 3%;
                    width: 3%;
                    margin-left: 48.5%;
                    margin-top: 2px;
                    margin-bottom: 2px;
                }
                </style>
        </head>
        <body>
            <div class="div1">
            
    
                    <div class="div2">
                        <div class="imgdiv">
                            <img  class="logo" src="https://drive.google.com/uc?authuser=0&id=1P0JiUm01B1itolWCpu06xmHilQcVtTFb&export=download" alt="logo">
                        </div>
                            <div class="div3">
                                <br>
                                    <p><span class="dot"></span>Your result is satisfactory!</p>
                                    <p><span class="dot"></span><b>But</b> remember <br>"A successful warrior is an average person with laser like focus."</p>
                                    <p><span class="dot"></span>Keep your focus straight.<br>We can see high potentiality in you.<br>
                                        We believe thet you can make upto the TOP 5 as well.</p>
                                        <p><span class="dot"></span>Remember that being average is too mainstream! So,</p>
                                        <p><span class="dot"></span>Wish you all the best for your upcoming tasks.<b>ROCK IT!!</b></p> 
                                        <p><span class="dot"></span>Your result for the month is as follows:</p>
                                        <img class="arrow" src="https://drive.google.com/uc?authuser=0&id=1V0u4yQmIJ9f1HTPMHH6rLGx04cycefw-&export=download" alt="down">
                                <br>
                            </div>
                    </div>
                </div>
            
        </body>
        </html>
        """

        body3 = """
        <html lang="en">
        <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <meta http-equiv="X-UA-Compatible" content="ie=edge">
        <title>Document</title>
        <style>
        .div1{
            border: 2px solid blue;
            
            margin: 5px;
        }
        .div2{
            margin: 10px;
            border: 2px solid black;
            border-radius: 20px;
        }
        .div3{
            margin: 10px;
            background-color: rgba(240,230,240,.9);
            border-radius: 20px;
            box-shadow: 0 2px 2px 0 rgb(0 ,0,0, .6 ) , 0 0 0 1px rgba(0,0,0,.3);
        }
        .imgdiv{
            padding-top: 2px;;
            /* float: left; */
            text-align: center;
        }
        .logo{
            height: 100%;
            width: 80%;
            box-shadow:0 2px 2px 0 rgb(0 ,0,0, .4 ) , 0 0 0 1px rgba(0,0,0,.2);
        }
    
        .dot{
            height: 10px;
            width: 10px;
            background-color: black;
            border-radius: 50%;
            display: inline-block;
            margin-right: 4px;
        }
        p{
            /* margin-left:20px; */
            text-align: center;
            font-size: 18px;
            font-family: 'Lucida Sans', 'Lucida Sans Regular', 'Lucida Grande', 'Lucida Sans Unicode', Geneva, Verdana, sans-serif;
        }
        .arrow{
            height: 3%;
            width: 3%;
            margin-left: 48.5%;
            margin-top: 2px;
            margin-bottom: 2px;
        }
        </style>
        </head>
        <body>
        <div class="div1">
            
    
            <div class="div2">
                <div class="imgdiv">
                    <img  class="logo" src="https://drive.google.com/uc?authuser=0&id=1P0JiUm01B1itolWCpu06xmHilQcVtTFb&export=download" alt="logo">
                </div>
                    <div class="div3">
                        <br>
                        <p><span class="dot"></span>You know how your performance was.</p>
                        <p><span class="dot"></span>Just don't let one bad grade define your true abilities.
                        <br>It is  just a minor setback.
                    <br>Don't get demotivated.</p>
                        <p><span class="dot"></span>How you choose to overcome it, is what sets you apart from the rest.</p>
                        <p><span class="dot"></span>We believe that you can do it like the rest.</p>
                        <p><span class="dot"></span>Give all your efforts</p>
                        <p><span class="dot"></span>Wish you all the best for your upcoming tasks. <b>ROCK IT!!</b></p>
                        <p><span class="dot"></span>Your result for the month is as follows:</p> 
                        <img class="arrow" src="https://drive.google.com/uc?authuser=0&id=1V0u4yQmIJ9f1HTPMHH6rLGx04cycefw-&export=download" alt="down">
                        <br>
    
                    </div>
            </div>
        </div>
        </body>
        </html>
        """
        if rank <= 5:
            msg.attach(MIMEText(body1, 'html'))
        elif (rank > 5) & (rank <= 15):
            msg.attach(MIMEText(body2, 'html'))
        else:
            msg.attach(MIMEText(body3, 'html'))

        filename = self.file_loc + "\\" + name + ".pdf"
        attachment = open(filename, 'rb')

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= " + filename)
        msg.attach(part)
        text = msg.as_string()
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_user, email_password)
        server.sendmail(email_user, email_send, text)

        server.quit()


Main_call = Main()
